import tkinter as tk
from tkinter import filedialog, scrolledtext
import ttkbootstrap as ttk
from ttkbootstrap.constants import *
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import locale
import threading
import os

# Tenta configurar o locale para português do Brasil de forma robusta
locales_para_tentar = ['pt-BR', 'pt_BR.utf8', 'Portuguese_Brazil', 'Portuguese']
for loc in locales_para_tentar:
    try:
        locale.setlocale(locale.LC_TIME, loc)
        break
    except locale.Error:
        continue

# --- CONFIGURAÇÃO ---
DIA_DE_CORTE = 5
nome_da_coluna_data = 'DATA'
nome_da_coluna_lote = 'LOTE'
nome_da_coluna_produto = 'QUEIJO'
nome_da_coluna_data_mat = 'DATA MAT'
nome_da_coluna_kg = 'KG'
produtos_desejados = ['CABOCLO', 'MATUTO', 'CORAÇÃO']

def ajustar_largura_colunas(sheet, max_widths=None):
    if max_widths is None:
        max_widths = {}
    for col in sheet.columns:
        max_length = 0
        column_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        if column_letter in max_widths and adjusted_width > max_widths[column_letter]:
            adjusted_width = max_widths[column_letter]
        sheet.column_dimensions[column_letter].width = adjusted_width

# --- FUNÇÃO DE PROCESSAMENTO FINAL E CORRIGIDA ---
def processar_planilha(caminho_arquivo_entrada, status_callback):
    try:
        nome_do_arquivo_saida = os.path.join(os.path.dirname(caminho_arquivo_entrada), 'Resumo_Queijos.xlsx')

        status_callback(f"Carregando o arquivo '{os.path.basename(caminho_arquivo_entrada)}'...")
        workbook = openpyxl.load_workbook(caminho_arquivo_entrada)
        
        status_callback("\nFASE 1: Consolidando o inventário...")
        inventario_total_df = pd.DataFrame()
        for nome_da_aba in workbook.sheetnames:
            if "Resumo_" in nome_da_aba: continue
            status_callback(f"Lendo dados da aba: '{nome_da_aba}'")
            df_aba = pd.read_excel(caminho_arquivo_entrada, sheet_name=nome_da_aba, skiprows=9)
            df_aba.columns = df_aba.columns.str.strip()
            df_aba[nome_da_coluna_data] = pd.to_datetime(df_aba[nome_da_coluna_data], errors='coerce')
            df_aba[nome_da_coluna_lote] = df_aba[nome_da_coluna_data].dt.dayofyear
            sheet = workbook[nome_da_aba]
            for index, row_data in df_aba.iterrows():
                if pd.notna(row_data.get(nome_da_coluna_lote)) and row_data.get(nome_da_coluna_produto) in produtos_desejados:
                    cell = sheet.cell(row=9 + index + 2, column=2); cell.value = int(row_data[nome_da_coluna_lote]); cell.number_format = '0'
            inventario_total_df = pd.concat([inventario_total_df, df_aba], ignore_index=True)

        inventario_total_df[nome_da_coluna_data_mat] = pd.to_datetime(inventario_total_df[nome_da_coluna_data_mat], errors='coerce')
        inventario_total_df.dropna(subset=[nome_da_coluna_data, nome_da_coluna_data_mat, nome_da_coluna_lote, nome_da_coluna_produto], inplace=True)
        
        status_callback("\nFASE 2: Processando o fluxo de estoque com a lógica correta...")
        resumos_mensais = {}
        meses_de_producao = sorted(inventario_total_df[nome_da_coluna_data].dt.to_period('M').unique())

        for periodo_producao in meses_de_producao:
            data_mes_producao = periodo_producao.to_timestamp()
            chave_mes = data_mes_producao.strftime('%B_%Y').capitalize()
            
            data_corte_final = (data_mes_producao + pd.DateOffset(months=1)).replace(day=DIA_DE_CORTE)
            status_callback(f"Analisando produção de {chave_mes.replace('_', ' ')} (Data de corte: {data_corte_final.strftime('%d/%m/%Y')})")
            
            inventario_ate_o_mes_df = inventario_total_df[inventario_total_df[nome_da_coluna_data].dt.to_period('M') <= periodo_producao]
            liberados_df = inventario_ate_o_mes_df[inventario_ate_o_mes_df[nome_da_coluna_data_mat] <= data_corte_final]
            pendentes_df = inventario_ate_o_mes_df[inventario_ate_o_mes_df[nome_da_coluna_data_mat] > data_corte_final]

            previsao_estoque_df = pd.DataFrame()
            if not pendentes_df.empty:
                df_temp = pendentes_df.copy()
                df_temp['MES_MATURACAO'] = df_temp[nome_da_coluna_data_mat].dt.strftime('%B/%Y').str.capitalize()
                previsao_estoque_df = df_temp.groupby(['MES_MATURACAO', nome_da_coluna_produto]).agg(TOTAL_KG_PENDENTE=(nome_da_coluna_kg, 'sum'), LOTES_PENDENTES=(nome_da_coluna_lote, lambda x: sorted(list(x)))).reset_index()
            
            produzido_no_mes_df = inventario_total_df[inventario_total_df[nome_da_coluna_data].dt.to_period('M') == periodo_producao]
            
            resumos_mensais[chave_mes] = {
                "lotes_liberados": liberados_df[nome_da_coluna_lote].astype(int).tolist(),
                "kg_liberados_por_queijo": liberados_df.groupby(nome_da_coluna_produto)[nome_da_coluna_kg].sum().to_dict(),
                "total_kg_produzidos": produzido_no_mes_df[nome_da_coluna_kg].sum(),
                "previsao_estoque": previsao_estoque_df
            }

        status_callback("\nFASE 3: Estilizando as abas de resumo...")
        
        font_titulo = Font(size=16, bold=True, color="FFFFFF", name='Calibri');fill_titulo = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid");align_center = Alignment(horizontal="center", vertical="center");font_secao = Font(size=13, bold=True, color="2F5496", name='Calibri');font_cabecalho_tabela = Font(size=11, bold=True, name='Calibri');align_left = Alignment(horizontal="left", vertical="center", wrap_text=True);align_right = Alignment(horizontal="right", vertical="center");number_format_kg = '#,##0.00';header_border = Border(bottom=Side(style='medium'));last_row_border = Border(top=Side(style='thin'));
        
        for chave_mes, dados in resumos_mensais.items():
            if not dados["lotes_liberados"] and not dados["total_kg_produzidos"]: continue
            nome_aba_resumo = f"Resumo_{chave_mes}";
            if nome_aba_resumo in workbook.sheetnames: workbook.remove(workbook[nome_aba_resumo])
            resumo_sheet = workbook.create_sheet(nome_aba_resumo); linha_atual = 1
            
            resumo_sheet.merge_cells(f'A{linha_atual}:D{linha_atual}'); titulo_cell = resumo_sheet[f'A{linha_atual}']; titulo_cell.value = f"Resumo da Produção - {chave_mes.replace('_', ' de ')}"; titulo_cell.font = font_titulo; titulo_cell.fill = fill_titulo; titulo_cell.alignment = align_center; resumo_sheet.row_dimensions[linha_atual].height = 30; linha_atual += 2
            
            resumo_sheet.cell(row=linha_atual, column=1, value="Produção do Mês").font = font_secao; linha_atual += 1
            resumo_sheet.cell(row=linha_atual, column=1, value="Total de KG Produzidos:").font = font_cabecalho_tabela; cell_kg_prod = resumo_sheet.cell(row=linha_atual, column=2, value=dados["total_kg_produzidos"]); cell_kg_prod.number_format = number_format_kg; cell_kg_prod.alignment = align_right; linha_atual += 2
            
            resumo_sheet.cell(row=linha_atual, column=1, value="Queijos Liberados para Venda").font = font_secao; linha_atual += 1
            resumo_sheet.cell(row=linha_atual, column=1, value="Lotes: " + ", ".join(map(str, sorted(dados["lotes_liberados"])))).alignment = align_left; resumo_sheet.merge_cells(f'A{linha_atual}:D{linha_atual}'); linha_atual += 1
            
            header_queijo = resumo_sheet.cell(row=linha_atual, column=1, value="QUEIJO"); header_queijo.font=font_cabecalho_tabela; header_queijo.border=header_border; header_queijo.alignment=align_left
            header_kg = resumo_sheet.cell(row=linha_atual, column=2, value="TOTAL KG LIBERADO"); header_kg.font=font_cabecalho_tabela; header_kg.border=header_border; header_kg.alignment=align_right
            linha_atual += 1
            for i, (queijo, kg_total) in enumerate(dados["kg_liberados_por_queijo"].items()):
                resumo_sheet.cell(row=linha_atual, column=1, value=queijo).alignment = align_left
                cell_k = resumo_sheet.cell(row=linha_atual, column=2, value=kg_total); cell_k.alignment = align_right; cell_k.number_format = number_format_kg
                linha_atual += 1
            resumo_sheet.cell(row=linha_atual, column=1).border = last_row_border; resumo_sheet.cell(row=linha_atual, column=2).border = last_row_border
            linha_atual += 2

            resumo_sheet.cell(row=linha_atual, column=1, value="Previsão de Estoque Pendente").font = font_secao; linha_atual += 1
            headers = ["Mês de Maturação", "Queijo", "Lotes", "Total KG Pendente"]; headers_align = [align_left, align_left, align_left, align_right]
            for col_num, header_title in enumerate(headers, 1):
                cell = resumo_sheet.cell(row=linha_atual, column=col_num, value=header_title); cell.font = font_cabecalho_tabela; cell.border = header_border; cell.alignment = headers_align[col_num-1]
            linha_atual += 1
            
            last_row_index = len(dados["previsao_estoque"]) -1
            for index, row in dados["previsao_estoque"].iterrows():
                lotes_str = ", ".join(map(str, row['LOTES_PENDENTES']))
                row_data = [row['MES_MATURACAO'], row[nome_da_coluna_produto], lotes_str, row['TOTAL_KG_PENDENTE']]
                data_aligns = [align_left, align_left, align_left, align_right]
                for col_num, value in enumerate(row_data, 1):
                    cell = resumo_sheet.cell(row=linha_atual, column=col_num, value=value); cell.alignment = data_aligns[col_num-1]
                resumo_sheet.cell(row=linha_atual, column=4).number_format = number_format_kg
                linha_atual += 1
            if last_row_index >= 0:
                for col_num in range(1, len(headers) + 1): resumo_sheet.cell(row=linha_atual, column=col_num).border = last_row_border

            ajustar_largura_colunas(resumo_sheet, max_widths={'C': 60})

        status_callback(f"\nSalvando o arquivo final...")
        workbook.save(nome_do_arquivo_saida)
        status_callback("\n--- PROCESSO CONCLUÍDO COM SUCESSO! ---")
        status_callback(f"\nO arquivo foi salvo na mesma pasta do original.")
    except Exception as e:
        status_callback(f"\n--- OCORREU UM ERRO! --- \nErro: {e}\nPor favor, verifique se o arquivo está correto e tente novamente.")

class App:
    def __init__(self, root):
        self.root = root;self.root.title("Analisador de Lotes de Queijo");self.root.geometry("800x600");self.filepath = tk.StringVar();self.frame = ttk.Frame(self.root, padding="10");self.frame.pack(fill=BOTH, expand=True);self.file_frame = ttk.Labelframe(self.frame, text="1. Selecione a Planilha", padding="10");self.file_frame.pack(fill=X, padx=5, pady=5);self.entry = ttk.Entry(self.file_frame, textvariable=self.filepath, state=DISABLED, width=70);self.entry.pack(side=LEFT, fill=X, expand=True, padx=(0, 5));self.browse_button = ttk.Button(self.file_frame, text="Procurar...", command=self.browse_file, style=SUCCESS);self.browse_button.pack(side=LEFT);self.run_button = ttk.Button(self.frame, text="2. Executar Análise", command=self.run_analysis, state=DISABLED, style=(SUCCESS, OUTLINE));self.run_button.pack(fill=X, padx=5, pady=10, ipady=10);self.status_box = scrolledtext.ScrolledText(self.frame, wrap=tk.WORD, height=20, font=("Consolas", 10));self.status_box.pack(fill=BOTH, expand=True, padx=5, pady=5);self.status_box.insert(tk.END, "Bem-vindo! Por favor, selecione a sua planilha Excel para começar.")
    def browse_file(self):
        path = filedialog.askopenfilename(title="Selecione a planilha Excel", filetypes=[("Arquivos Excel", "*.xlsx")]);
        if path:self.filepath.set(path);self.run_button.config(state=NORMAL);self.status_box.delete('1.0', tk.END);self.status_box.insert(tk.END, f"Arquivo selecionado: {path}\n\nPronto para iniciar a análise.")
    def update_status(self, message):
        self.status_box.insert(tk.END, message + "\n");self.status_box.see(tk.END);self.root.update_idletasks()
    def run_analysis(self):
        path = self.filepath.get();
        if not path:self.update_status("ERRO: Nenhum arquivo selecionado.");return
        self.run_button.config(state=DISABLED);self.browse_button.config(state=DISABLED);self.status_box.delete('1.0', tk.END);analysis_thread = threading.Thread(target=processar_planilha, args=(path, self.update_status));analysis_thread.start();self.monitor_thread(analysis_thread)
    def monitor_thread(self, thread):
        if thread.is_alive():self.root.after(100, lambda: self.monitor_thread(thread))
        else:self.run_button.config(state=NORMAL);self.browse_button.config(state=NORMAL)

if __name__ == "__main__":
    root = ttk.Window(themename="flatly")
    app = App(root)
    root.mainloop()